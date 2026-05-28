from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from tests.test_classes_students_courses_schedules import auth_headers


def workbook_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["student_no", "name", "gender", "remark"])
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def malformed_xlsx_zip_bytes():
    stream = BytesIO()
    with ZipFile(stream, "w") as archive:
        archive.writestr("dummy.txt", "not an Excel workbook")
    stream.seek(0)
    return stream


def test_student_import_template_downloads_expected_workbook(client):
    headers = auth_headers(client, "teacher1")

    response = client.get("/api/v1/students/import-template", headers=headers)

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["content-disposition"] == 'attachment; filename="student-import-template.xlsx"'
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet.title == "students"
    assert [cell.value for cell in sheet[1]] == ["student_no", "name", "gender", "status", "remark"]


def test_student_import_writes_valid_rows_and_records_errors(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    file_obj = workbook_bytes([["S001", "张三", "男", ""], ["S001", "重复", "女", ""]])

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                file_obj,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["success_count"] == 1
    assert body["failed_count"] == 1
    assert body["status"] == "partial_success"

    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers)
    assert errors.json()["items"][0]["row_number"] == 3
    assert errors.json()["items"][0]["field"] == "student_no"


def test_student_import_can_update_existing_student(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    client.post(
        "/api/v1/students",
        headers=headers,
        json={"class_id": class_id, "student_no": "S001", "name": "旧姓名", "gender": "男"},
    )
    file_obj = workbook_bytes([["S001", "新姓名", "女", "updated"]])

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}&update_existing=true",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                file_obj,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "success"
    students = client.get("/api/v1/students", headers=headers).json()["items"]
    assert students[0]["name"] == "新姓名"
    assert students[0]["gender"] == "女"
    assert students[0]["remark"] == "updated"


def test_student_import_duplicate_seen_even_when_first_row_invalid(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    file_obj = workbook_bytes([["S001", "", "男", ""], ["S001", "张三", "男", ""]])

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                file_obj,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["success_count"] == 0
    assert body["failed_count"] == 2
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers).json()
    assert [item["field"] for item in errors["items"]] == ["name", "student_no"]
    assert errors["items"][1]["reason"] == "同一文件中学号重复"


def test_student_import_unowned_class_records_failed_batch_without_target_display(client):
    owner_headers = auth_headers(client, "teacher1")
    other_headers = auth_headers(client, "teacher2")
    other_class_id = client.post(
        "/api/v1/classes",
        headers=other_headers,
        json={"name": "二班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    file_obj = workbook_bytes([["S001", "张三", "男", ""]])

    response = client.post(
        f"/api/v1/students/import?target_class_id={other_class_id}",
        headers=owner_headers,
        files={
            "file": (
                "students.xlsx",
                file_obj,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["success_count"] == 0
    assert body["failed_count"] == 1
    assert body["status"] == "failed"
    batch = client.get(f"/api/v1/imports/{body['batch_id']}", headers=owner_headers).json()
    assert batch["target_class_id"] is None
    assert batch["target_class_name"] is None
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=owner_headers).json()
    assert errors["items"][0]["field"] == "target_class_id"


def test_student_import_corrupt_upload_returns_failed_batch(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                BytesIO(b"not an excel workbook"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "failed"
    assert body["success_count"] == 0
    assert body["failed_count"] == 1
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers).json()
    assert errors["items"][0]["field"] == "file"


def test_student_import_malformed_xlsx_zip_returns_failed_batch(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                malformed_xlsx_zip_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "failed"
    assert body["success_count"] == 0
    assert body["failed_count"] == 1
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers).json()
    assert errors["items"][0]["field"] == "file"


def test_student_import_wrong_extension_returns_failed_batch(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={"file": ("students.txt", BytesIO("student_no,name\nS001,张三".encode()), "text/plain")},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "failed"
    assert body["success_count"] == 0
    assert body["failed_count"] == 1
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers).json()
    assert errors["items"][0]["field"] == "file"


def test_student_import_invalid_status_records_row_error(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    wb = Workbook()
    ws = wb.active
    ws.append(["student_no", "name", "gender", "status", "remark"])
    ws.append(["S001", "张三", "男", "deleted", ""])
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                file_obj,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "failed"
    assert body["success_count"] == 0
    assert body["failed_count"] == 1
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers).json()
    assert errors["items"][0]["row_number"] == 2
    assert errors["items"][0]["field"] == "status"


def test_student_import_clamps_overlong_upload_filename(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    long_filename = f"{'students-' * 40}.xlsx"

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                long_filename,
                workbook_bytes([["S001", "张三", "男", ""]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    batch = client.get(f"/api/v1/imports/{body['batch_id']}", headers=headers).json()
    assert len(batch["file_name"]) <= 255


def test_student_import_clamps_overlong_raw_error_payload(client):
    headers = auth_headers(client, "teacher1")
    class_id = client.post(
        "/api/v1/classes",
        headers=headers,
        json={"name": "一班", "grade": "七年级", "academic_year": "2026-2027"},
    ).json()["id"]
    long_student_no = "S" * 20_000

    response = client.post(
        f"/api/v1/students/import?target_class_id={class_id}",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                workbook_bytes([[long_student_no, "张三", "男", ""]]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "failed"
    errors = client.get(f"/api/v1/imports/{body['batch_id']}/errors", headers=headers).json()
    error = errors["items"][0]
    assert error["field"] == "student_no"
    assert len(error["raw_value"]) <= 10_000
    assert len(error["raw_data"]) <= 10_000
